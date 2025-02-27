"""## REMOVENDO OS ESPAÇOS NA REFERÊNCIA"""

import pandas as pd
from openpyxl import load_workbook

def remove_leading_trailing_spaces():
    # Caminho do arquivo de entrada e saída
    input_file = "ARTES VISUAIS.xlsx"
    output_file = "ARTES_VISUAIS_ATUALIZADO.xlsx"
    sheet_name = "RelatorioCursoBibliografia"
    column_name = "BIBLIOGRAFIA"

    # Carregar a planilha mantendo a formatação
    wb = load_workbook(input_file)
    ws = wb[sheet_name]

    # Identificar a coluna correta
    col_idx = None
    for col in ws.iter_cols(min_row=1, max_row=1):
        if col[0].value == column_name:
            col_idx = col[0].column
            break

    if col_idx is None:
        print(f"Coluna '{column_name}' não encontrada.")
        return

    # Remover espaços no início e no final de cada célula na coluna especificada
    for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
        for c in cell:
            if isinstance(c.value, str):
                c.value = c.value.strip()

    # Salvar o arquivo sem alterar a formatação
    wb.save(output_file)
    print(f"Arquivo salvo em: {output_file}")

# Executar a função
remove_leading_trailing_spaces()
