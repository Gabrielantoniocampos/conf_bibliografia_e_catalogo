### PADRONIZAR A COLUNA DE REFERÊNCIA BIBLIOGRÁFICA CONFORME A ABNT, REMOVENDO AS STRINGS DESNECESSÁRIAS

import pandas as pd

# Carregar a planilha Excel para um DataFrame
df = pd.read_excel("/content/Catálogo MB (1).xlsx")

# Função para tratar a coluna "ABNT"
def tratar_abnt(texto):
    # Garantir que o texto seja uma string antes de realizar substituições
    texto = str(texto).replace("<p>", "")  # Remove a tag <p> do texto
    
    # Verificar se a palavra "Ebook" está presente no texto
    if "Ebook" in texto:
        # Remover tudo a partir da palavra "Ebook"
        texto = texto.split("Ebook")[0]
    
    # Retornar o texto tratado, removendo espaços em branco no início e no final
    return texto.strip()

# Aplicar a função `tratar_abnt` a todos os elementos da coluna "ABNT"
df["ABNT"] = df["ABNT"].apply(tratar_abnt)

# Salvar o DataFrame tratado em um novo arquivo Excel
df.to_excel("Catálogo MB Tratado.xlsx", index=False)

### REMOVAR LINHAS DA PLANILHA IMPORTADA DO SAG COM A BIBLIOGRAFIA DO CURSO

from openpyxl import load_workbook

def remove_leading_trailing_spaces():
    """
    Função para remover espaços em branco no início e no final das células de uma coluna específica
    em uma planilha Excel, mantendo a formatação original do arquivo.
    """
    # Caminho do arquivo de entrada e saída
    input_file = "ARTES VISUAIS.xlsx"  # Arquivo Excel de entrada
    output_file = "ARTES_VISUAIS_ATUALIZADO.xlsx"  # Arquivo Excel de saída
    sheet_name = "RelatorioCursoBibliografia"  # Nome da planilha a ser processada
    column_name = "BIBLIOGRAFIA"  # Nome da coluna a ser tratada
    
    # Carregar a planilha mantendo a formatação usando openpyxl
    wb = load_workbook(input_file)  # Carrega o arquivo Excel
    ws = wb[sheet_name]  # Seleciona a planilha específica
    
    # Identificar a coluna correta pelo nome no cabeçalho (primeira linha)
    col_idx = None
    for col in ws.iter_cols(min_row=1, max_row=1):  # Itera sobre as colunas da primeira linha
        if col[0].value == column_name:  # Verifica se o nome da coluna corresponde
            col_idx = col[0].column  # Armazena o índice da coluna
            break
    
    # Verificar se a coluna foi encontrada
    if col_idx is None:
        print(f"Coluna '{column_name}' não encontrada.")
        return
    
    # Remover espaços no início e no final de cada célula na coluna especificada
    for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):  # Itera sobre as células da coluna
        for c in cell:
            if isinstance(c.value, str):  # Verifica se o valor da célula é uma string
                c.value = c.value.strip()  # Remove espaços em branco no início e no final
    
    # Salvar o arquivo sem alterar a formatação
    wb.save(output_file)  # Salva o arquivo tratado
    print(f"Arquivo salvo em: {output_file}")

# Executar a função
remove_leading_trailing_spaces()
