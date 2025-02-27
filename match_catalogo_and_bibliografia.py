from openpyxl.styles import PatternFill
from openpyxl import load_workbook

def highlight_matching_cells():
    """
    Função para destacar células em uma planilha Excel que contêm valores correspondentes
    a uma coluna específica de outra planilha.
    """
    # Caminhos dos arquivos
    file_artes = "/content/ARTES_VISUAIS_ATUALIZADO.xlsx"  # Arquivo de Artes Visuais
    file_catalogo = "/content/Catálogo MB Tratado.xlsx"  # Arquivo do Catálogo MB
    output_file = "ARTES_VISUAIS_DESTACADO.xlsx"  # Arquivo de saída com células destacadas

    # Nome das planilhas
    sheet_artes = "RelatorioCursoBibliografia"  # Planilha de Artes Visuais
    sheet_catalogo = "Sheet1"  # Planilha do Catálogo MB

    # Carregar os arquivos Excel mantendo a formatação
    wb_artes = load_workbook(file_artes)  # Carrega o arquivo de Artes Visuais
    ws_artes = wb_artes[sheet_artes]  # Seleciona a planilha de Artes Visuais

    wb_catalogo = load_workbook(file_catalogo)  # Carrega o arquivo do Catálogo MB
    ws_catalogo = wb_catalogo[sheet_catalogo]  # Seleciona a planilha do Catálogo MB

    # Coletar os valores da coluna R (coluna 18) da planilha Catálogo MB
    valores_catalogo = set()  # Usamos um conjunto para armazenar valores únicos
    for row in ws_catalogo.iter_rows(min_row=2, min_col=18, max_col=18, values_only=True):
        if row[0] is not None:  # Verifica se a célula não está vazia
            valores_catalogo.add(str(row[0]).strip())  # Adiciona o valor tratado ao conjunto

    # Definir o preenchimento verde para destacar células
    fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Percorrer a coluna N (coluna 14) da planilha Artes Visuais e aplicar a formatação
    for row in ws_artes.iter_rows(min_row=2, min_col=14, max_col=14):
        cell = row[0]  # Célula atual na coluna N
        if cell.value and str(cell.value).strip() in valores_catalogo:  # Verifica se o valor corresponde
            cell.fill = fill_green  # Aplica o preenchimento verde

    # Salvar o arquivo atualizado
    wb_artes.save(output_file)
    print(f"Arquivo salvo em: {output_file}")

# Executar a função
highlight_matching_cells()
